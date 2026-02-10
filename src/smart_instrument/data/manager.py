import csv
from datetime import datetime
from openpyxl import Workbook
import os
import subprocess

class DataManager:
    def __init__(self):
        # 数据存储
        self.data = []
        self.column_count = 0
        # Excel相关
        self.excel_workbook = None
        self.excel_sheet = None
        self.excel_filename = None
        self.excel_initialized = False
    
    def initialize_excel(self):
        """初始化Excel文件"""
        try:
            # 创建新的Excel工作簿
            self.excel_workbook = Workbook()
            self.excel_sheet = self.excel_workbook.active
            self.excel_sheet.title = "测试数据"
            
            # 设置标题行
            self.excel_sheet.cell(row=1, column=1, value="设备")
            self.excel_sheet.cell(row=2, column=1, value="IT8811 (电阻)")
            self.excel_sheet.cell(row=3, column=1, value="DMM6500 (电压)")
            self.excel_sheet.cell(row=4, column=1, value="KEYSIGHT 34461A (电流)")
            
            # 生成文件名
            self.excel_filename = f"test_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # 保存Excel文件
            self.excel_workbook.save(self.excel_filename)
            
            # 打开Excel文件
            try:
                if os.name == 'nt':  # Windows
                    os.startfile(self.excel_filename)
                elif os.name == 'posix':  # macOS
                    subprocess.run(['open', self.excel_filename], check=True)
                else:  # Linux
                    subprocess.run(['xdg-open', self.excel_filename], check=True)
            except Exception as e:
                print(f"打开Excel文件失败: {str(e)}")
            
            self.excel_initialized = True
            return True, f"Excel文件已创建并打开: {self.excel_filename}"
        except Exception as e:
            return False, f"初始化Excel失败: {str(e)}"
    
    def record_data(self, resistance, voltage, current=None):
        """记录数据"""
        try:
            # 第一次点击时初始化Excel
            if not self.excel_initialized:
                excel_success, excel_msg = self.initialize_excel()
                if not excel_success:
                    print(f"Excel初始化失败: {excel_msg}")
            
            # 更新列计数
            self.column_count += 1
            
            # 保存到数据列表
            if len(self.data) < 3:
                self.data.append([])  # IT8811数据
                self.data.append([])  # DMM6500数据
                self.data.append([])  # KEYSIGHT 34461A数据
            
            self.data[0].append(resistance)
            self.data[1].append(voltage)
            self.data[2].append(current if current else "")
            
            # 实时写入Excel
            if self.excel_workbook and self.excel_sheet:
                try:
                    # 添加触发列标题
                    self.excel_sheet.cell(row=1, column=self.column_count + 1, value=f"触发{self.column_count}")
                    
                    # 写入数据
                    self.excel_sheet.cell(row=2, column=self.column_count + 1, value=resistance)
                    self.excel_sheet.cell(row=3, column=self.column_count + 1, value=voltage)
                    self.excel_sheet.cell(row=4, column=self.column_count + 1, value=current if current else "")
                    
                    # 保存Excel文件
                    self.excel_workbook.save(self.excel_filename)
                except Exception as e:
                    print(f"写入Excel失败: {str(e)}")
            
            return True, self.column_count
        except Exception as e:
            return False, f"记录数据失败: {str(e)}"
    
    def save_to_csv(self):
        """保存数据到CSV文件"""
        if not self.data:
            return False, "没有数据可保存"
        
        try:
            filename = f"test_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            with open(filename, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                
                # 写入标题
                headers = ['设备']
                for i in range(1, self.column_count + 1):
                    headers.append(f'触发{i}')
                writer.writerow(headers)
                
                # 写入IT8811数据
                row1 = ['IT8811 (电阻)'] + self.data[0]
                writer.writerow(row1)
                
                # 写入DMM6500数据
                row2 = ['DMM6500 (电压)'] + self.data[1]
                writer.writerow(row2)
                
                # 写入KEYSIGHT 34461A数据
                row3 = ['KEYSIGHT 34461A (电流)'] + (self.data[2] if len(self.data) > 2 else [])
                writer.writerow(row3)
            
            return True, f"数据已保存到 {filename}"
        except Exception as e:
            return False, f"保存数据失败: {str(e)}"
    
    def clear_data(self):
        """清空数据"""
        self.data = []
        self.column_count = 0
        return True, "数据已清空"
